import os
import json
import cv2
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import re
from matplotlib.animation import FFMpegWriter
from matplotlib.gridspec import GridSpec
from matplotlib.colors import to_rgba
import time
from tqdm import tqdm
from threading import Thread
from queue import Queue, Empty
from openpyxl.styles import PatternFill, Font

MAKE_VIDEO = False  # ← 여기만 True/False로 바꿔 쓰면 됨

# ----------------------------
# 글로벌 폰트 설정 (전체 폰트 크기 축소)
# ----------------------------
plt.rcParams.update({
    'font.size': 8,              # 기본 폰트 크기
    'axes.titlesize': 8,         # 서브플롯 제목
    'axes.labelsize': 8,         # 축 레이블
    'xtick.labelsize': 8,        # x축 눈금 레이블
    'ytick.labelsize': 8,        # y축 눈금 레이블
    'legend.fontsize': 8,        # 범례
    'figure.titlesize': 10       # 전체 figure 제목 (없을 경우 무시)
})

# ----------------------------
# 설정
# ----------------------------
FPS = 15
FRAME_WIDTH = 228
FRAME_HEIGHT = 128
WINDOW_SECONDS = 60
STEP_FRAMES = WINDOW_SECONDS * FPS
OUTPUT_NAME = "test.mp4"
SHADING_FREQ = FPS  # 초당 한 번만 음영 업데이트

# ----------------------------
# FrameLoader: 비디오 프레임 프리패칭
# ----------------------------
class FrameLoader:
    def __init__(self, cap, maxsize=30):
        self.cap = cap
        self.q = Queue(maxsize=maxsize)
        self.stopped = False
        self.thread = Thread(target=self._load, daemon=True)
        self.thread.start()
    def _load(self):
        while not self.stopped:
            ret, frame = self.cap.read()
            if not ret:
                self.stop()
                break
            try:
                self.q.put(frame, timeout=1)
            except:
                pass
    def read(self, timeout=1):
        try:
            return self.q.get(timeout=timeout)
        except Empty:
            return None
    def stop(self):
        self.stopped = True
        try:
            while True:
                self.q.get(False)
        except Empty:
            pass

# ----------------------------
# 타임라인 데이터 로드
# ----------------------------
def open_timeline_data(timeline_dir, config_path, start_frame):
    start = time.time()
    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)
    with open(os.path.join(timeline_dir, 'global_stats.json'), 'r', encoding='utf-8') as f:
        global_stats = json.load(f)
    data_dict, caps = {}, {}
    total_frames = np.inf
    for fname in sorted(os.listdir(timeline_dir)):
        if not fname.endswith('_augmented.csv'): continue
        pid = fname.replace('_augmented.csv','').split('_')[-1]
        df = pd.read_csv(os.path.join(timeline_dir, fname))
        data_dict[pid] = df
        total_frames = min(total_frames, len(df))

        if MAKE_VIDEO:
            base = fname.replace('_augmented.csv','')
            for ext in ('.mp4', '.avi'):
                path = os.path.join(timeline_dir, base + ext)
                if os.path.isfile(path):
                    cap = cv2.VideoCapture(path)
                    cap.set(cv2.CAP_PROP_POS_FRAMES, start_frame)
                    caps[pid] = FrameLoader(cap)
                    break
    print(f"[TIME] Data loading: {time.time() - start:.2f}s")
    return config, global_stats, data_dict, caps, int(total_frames)

# ----------------------------
# nod 이벤트 탐지 (펄스 기반)
# ----------------------------
def detect_nod_events(arr, fall_thresh, rise_thresh, max_dur, min_cyc, fps):
    # stop_thresh 파라미터는 더 이상 사용하지 않음
    mask = np.zeros_like(arr, dtype=int)
    state = 'idle'
    cycle_count = 0
    event_start = None
    max_frames = int(max_dur * fps)
    for i, v in enumerate(arr):
        if state == 'idle':
            if v > fall_thresh:
                state = 'down'
                event_start = i
        elif state == 'down':
            if v < rise_thresh:
                cycle_count += 1
                if cycle_count >= min_cyc and event_start is not None and (i - event_start) <= max_frames:
                    mask[i] = 1  # 주기 완료 시점에 펄스 마킹
                    state = 'idle'
                    cycle_count = 0
                    event_start = None
        # 최대 지속시간 초과 시 초기화
        if event_start is not None and (i - event_start) > max_frames:
            state = 'idle'
            cycle_count = 0
            event_start = None
    return mask

# ----------------------------
# 동시성 마스크 계산
# ----------------------------
def calculate_synchrony_mask(data_dict, config, global_stats):
    start = time.time()
    masks = {}
    for name, cfg in config.items():
        ctype = cfg.get('type')
        win = int(cfg.get('sync_window', 0.3) * FPS)

        # Numeric & Categorical
        if ctype in ('numeric', 'categorical'):
            thr = cfg.get('threshold_std', 2.0)
            mode = cfg.get('sync_direction', 'same')
            mats_above, mats_below = [], []
            for pid, df in data_dict.items():
                if ctype == 'numeric':
                    raw = df[cfg['column']].astype(float).values
                else:
                    raw = df[cfg['column']].fillna('neutral') \
                          .map(cfg['mapping']).fillna(0).astype(float).values
                succ = (df['success'].astype(int).values == 1) if 'success' in df.columns else np.ones_like(raw, bool)
                if cfg.get('zscore', False):
                    m, s = global_stats[pid][name]['mean'], global_stats[pid][name]['std']
                    raw = (raw - m) / s
                mats_above.append((raw > thr) & succ)
                mats_below.append((raw < -thr) & succ)
            mats_above = np.vstack(mats_above)
            mats_below = np.vstack(mats_below)
            ext_above = np.zeros_like(mats_above, dtype=bool)
            ext_below = np.zeros_like(mats_below, dtype=bool)
            for idx in range(mats_above.shape[0]):
                r_ab, r_bl = mats_above[idx], mats_below[idx]
                ext_r_ab = np.zeros_like(r_ab)
                ext_r_bl = np.zeros_like(r_bl)
                for t in range(len(r_ab)):
                    start_t = max(0, t - win)
                    if r_ab[start_t:t+1].any(): ext_r_ab[t] = True
                    if r_bl[start_t:t+1].any(): ext_r_bl[t] = True
                ext_above[idx] = ext_r_ab
                ext_below[idx] = ext_r_bl
            if mode == 'any':
                # sync = ext_above.sum(axis=0) + ext_below.sum(axis=0)
                sync = np.logical_or(ext_above, ext_below).sum(axis=0)
            elif mode == 'same':
                sync = np.maximum(ext_above.sum(axis=0), ext_below.sum(axis=0))
            elif mode == 'positive':
                sync = ext_above.sum(axis=0)
            elif mode == 'negative':
                sync = ext_below.sum(axis=0)
            else:
                sync = np.zeros_like(ext_above.sum(axis=0), dtype=int)
            masks[name] = sync

        # Event
        elif ctype == 'event':
            params = cfg['event_params']
            fall = params.get('fall_z_thresh', 1.0)
            rise = params.get('rise_z_thresh', -1.0)
            md   = params.get('max_duration', 0.6)
            mc   = params.get('min_cycles', 1)
            mats = []
            for pid, df in data_dict.items():
                raw = df[cfg['column']].astype(float).values
                succ = (df['success'].astype(int).values == 1) if 'success' in df.columns else np.ones_like(raw, bool)
                if cfg.get('zscore', False):
                    m, s = global_stats[pid][name]['mean'], global_stats[pid][name]['std']
                    raw = (raw - m) / s
                ev = detect_nod_events(raw, fall, rise, md, mc, FPS).astype(bool)
                # ★ success 적용
                ev = ev & succ
                mats.append(ev)
            mats = np.vstack(mats)
            ext = np.zeros_like(mats, dtype=bool)
            for idx in range(mats.shape[0]):
                r = mats[idx]
                ext_r = np.zeros_like(r)
                for t in range(len(r)):
                    start_t = max(0, t - win)
                    if r[start_t:t+1].any(): ext_r[t] = True
                ext[idx] = ext_r
            masks[name] = ext.sum(axis=0)
        else:
            continue

    print(f"[TIME] Synchrony mask calc: {time.time()-start:.2f}s")
    return masks

# ----------------------------
# 메인 시각화
# ----------------------------
def visualize_timeline_optimized(timeline_dir, config_path, start_time=None, end_time=None):
    total_start = time.time()
    sf = 0 if start_time is None else int(start_time * FPS)

    config, global_stats, data_dict, caps, total_frames = open_timeline_data(timeline_dir, config_path, sf)
    ef = min(int(end_time * FPS), total_frames) if end_time else total_frames
    if not data_dict:
        print('[SKIP] No data')
        return
    
    pids = list(data_dict.keys())
    indicators = list(config.items())
    colors = plt.cm.tab10.colors

    sync_masks = calculate_synchrony_mask(data_dict, config, global_stats)

    # 사용자가 직접 지정할 하이라이트 지표 리스트 (원하는 지표명을 추가)
    HIGHLIGHT_INDICATORS = [
        "face_distance"
    ]

    # ----------------------------
    # 동시성 카운트 결과 CSV 저장 (wide format)
    # ----------------------------
    def _sanitize_sheet_name(name: str) -> str:
        return re.sub(r'[:\\/?*\[\]]', '_', str(name))[:31]

    sync_xlsx = os.path.join(timeline_dir, 'sync_counts.xlsx')
    max_p = len(pids)

    # --- [추가] 경로로부터 메타 정보 추출 ---
    parts = os.path.normpath(timeline_dir).split(os.sep)
    semester_raw = parts[-4]   # "24-1"
    group = parts[-3]          # "A4"
    week = parts[-2]           # "W1"
    timeline_idx = parts[-1]   # "T1"
    year_prefix, sem_num = semester_raw.split("-")  # "24", "1"
    year = "20" + year_prefix  # "2024"
    semester = sem_num         # "1"

    with pd.ExcelWriter(sync_xlsx, engine='openpyxl') as writer:
        # --- [추가] 1) Meta 시트를 '가장 먼저' 기록해서 엑셀 첫 탭이 되도록 ---
        df_meta = pd.DataFrame([{
            "year": year,
            "semester": semester,
            "group": group,
            "week": week,
            "timeline": timeline_idx,
            "total_participants": max_p,
            "total_frames": total_frames
        }])
        df_meta.to_excel(writer, sheet_name="Meta", index=False)

        # --- 2) 기존 지표별 동시성 카운트 시트들 (기능 유지) ---
        for ind_name, mask in sync_masks.items():
            # 0~max_p 동시 인원수 카운트 (기존 로직 유지)
            counts = [int((mask == i).sum()) for i in range(max_p + 1)]

            # wide 테이블 (기존 로직 유지)
            df_one = pd.DataFrame([counts], columns=[str(i) for i in range(max_p + 1)])
            df_one.index = [ind_name]
            df_one.index.name = "indicator"

            sheet_name = _sanitize_sheet_name(ind_name)
            df_one.to_excel(writer, sheet_name=sheet_name, index=True)

            # 하이라이트 (기존 로직 유지)
            if ind_name in HIGHLIGHT_INDICATORS:
                ws = writer.sheets[sheet_name]
                name_cell = ws.cell(row=2, column=1)  # A2
                name_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                name_cell.font = Font(bold=True)

    print(f"[완료] 동시성 결과 저장 → {sync_xlsx}")

    # ----------------------------
    # 프레임별 개별 행동 마스크 저장 (원시 -1/0/1 → Excel + 분석 시트)
    # ----------------------------
    mask_xlsx = os.path.join(timeline_dir, 'sync_mask.xlsx')
    mask_dict = {}

    # 1) 원시 -1/0/1 마스크 생성 (윈도 확장 X, 방향/임계만 반영, success 적용 O)
    for name, cfg in config.items():
        ctype = cfg.get('type')
        thr = cfg.get('threshold_std', 2.0)

        for pid, df in data_dict.items():
            # success==1 프레임만 유효 (없으면 전체 True)
            succ = (df['success'].astype(int).values == 1) if 'success' in df.columns else np.ones(len(df), dtype=bool)

            if ctype in ('numeric', 'categorical'):
                if ctype == 'numeric':
                    raw = df[cfg['column']].astype(float).values
                else:
                    # categorical은 수치 매핑으로 받아 ±임계 비교 (bool 캐스팅 금지)
                    raw = df[cfg['column']].fillna('neutral').map(cfg['mapping']).astype(float).values

                if cfg.get('zscore', False):
                    m = global_stats[pid][name]['mean']
                    s = global_stats[pid][name]['std'] or 0.0
                    raw = (raw - m) / (s if s != 0 else 1.0)

                above = raw >  thr
                below = raw < -thr

                signed = np.zeros_like(raw, dtype=int)
                signed[above & succ] =  1
                signed[below & succ] = -1
                # 나머지는 0 (success==0 포함)
                mask_dict[(name, pid)] = signed

            elif ctype == 'event':
                vals = df[cfg['column']].astype(float).values
                if cfg.get('zscore', False):
                    m = global_stats[pid][name]['mean']
                    s = global_stats[pid][name]['std'] or 0.0
                    vals = (vals - m) / (s if s != 0 else 1.0)

                ep = cfg.get('event_params', {})
                ev = detect_nod_events(
                    vals,
                    ep.get('fall_z_thresh', 2.0),
                    ep.get('rise_z_thresh', -1.0),
                    ep.get('max_duration', 0.5),
                    ep.get('min_cycles', 2),
                    FPS
                ).astype(bool)

                signed = np.zeros_like(vals, dtype=int)
                signed[ev & succ] = 1   # 이벤트는 부호 개념 없음 → 1/0
                mask_dict[(name, pid)] = signed

            else:
                continue

    mask_df = pd.DataFrame(mask_dict)
    mask_df.index.name = 'frame'
    mask_df.columns = pd.MultiIndex.from_tuples(mask_df.columns, names=['indicator','pid'])

    # ---------- (A) 지표/참가자 표시 순서 고정 ----------
    indicator_order = list(config.keys())      # config 정의 순서
    pid_order       = list(data_dict.keys())   # 로딩된 참가자 순서
    # 존재하는 컬럼만 유지하여 재인덱싱
    from itertools import product
    desired_cols = [c for c in product(indicator_order, pid_order) if c in set(mask_df.columns)]
    mask_df = mask_df.reindex(columns=pd.MultiIndex.from_tuples(desired_cols, names=['indicator','pid']), copy=False)

    # 2) calculate_synchrony_mask와 동일한 규칙으로 재집계 (윈도 확장 + 방향 모드, any=합집합)
    def _extend_bool(b: np.ndarray, win: int) -> np.ndarray:
        """과거 win 프레임 동안 하나라도 True면 현재 True"""
        if win <= 0:
            return b.astype(bool)
        out = np.zeros_like(b, dtype=bool)
        n = len(b)
        for t in range(n):
            s = max(0, t - win)
            if b[s:t+1].any():
                out[t] = True
        return out

    perframe_cols = {}   # key=(indicator, level_str) → 0/1 Series(index=frame)
    counts_series = {}   # key=indicator → pd.Series(level -> count)

    P = len(pid_order)   # 참가자 수(모든 모드에서 최대 P가 상한)
    for ind in indicator_order:
        sub = mask_df[ind]                 # frame × pid, 값 ∈ {-1,0,1}
        above_raw = (sub ==  1).to_numpy() # F×P
        below_raw = (sub == -1).to_numpy() # F×P

        win = int(config[ind].get('sync_window', 0.3) * FPS)

        # per-pid 윈도 확장
        ext_above = np.vstack([_extend_bool(above_raw[:, j], win) for j in range(P)]).T  # F×P
        ext_below = np.vstack([_extend_bool(below_raw[:, j], win) for j in range(P)]).T  # F×P

        pos_levels  = ext_above.sum(axis=1)                 # 0..P
        neg_levels  = ext_below.sum(axis=1)                 # 0..P
        mode = config[ind].get('sync_direction', 'same')

        if mode == 'positive':
            levels = pos_levels
        elif mode == 'negative':
            levels = neg_levels
        elif mode == 'same':
            levels = np.maximum(pos_levels, neg_levels)
        elif mode == 'any':
            # ★ any = 합집합(OR) → 0..P
            levels = np.logical_or(ext_above, ext_below).sum(axis=1)
        else:
            levels = np.zeros_like(pos_levels, dtype=int)

        # PerFrameLevels: (ind, '0'..'P') 원-핫(0/1)
        levels_s = pd.Series(levels, index=sub.index)
        for k in range(0, P + 1):
            perframe_cols[(ind, str(k))] = (levels_s == k).astype(int)

        # LevelCounts: 0..P 프레임 수
        vc = levels_s.value_counts().reindex(range(0, P + 1), fill_value=0)
        vc.index = vc.index.astype(str)
        counts_series[ind] = vc

    # PerFrameLevels DF (지표/레벨 순서 고정: indicator_order × 0..P)
    level_cols = [str(k) for k in range(0, P + 1)]
    perframe_df = pd.DataFrame(perframe_cols, index=mask_df.index)
    perframe_df.columns = pd.MultiIndex.from_tuples(perframe_df.columns, names=['indicator', 'level'])
    perframe_df = perframe_df.reindex(
        columns=pd.MultiIndex.from_product([indicator_order, level_cols]),
        copy=False
    )
    perframe_df.index.name = 'frame'

    # LevelCounts DF (지표 순서 고정)
    def _flat_label(x):
        return x if isinstance(x, str) else "_".join(map(str, x)) if isinstance(x, tuple) else str(x)

    indicator_labels = [_flat_label(x) for x in indicator_order]
    counts_df = pd.DataFrame(
        0,
        index=pd.Index(indicator_labels, name='indicator'),
        columns=level_cols,
        dtype=int
    )
    for ind, ser in counts_series.items():
        counts_df.loc[_flat_label(ind), ser.index.astype(str)] = ser.values

    # 3) 엑셀로 저장 (세 시트 모두 동일 지표 순서)
    with pd.ExcelWriter(mask_xlsx, engine='openpyxl') as writer:
        mask_df.to_excel(writer, sheet_name='RawMask')            # -1/0/1 원시 마스크
        perframe_df.to_excel(writer, sheet_name='PerFrameLevels') # 프레임×(지표,수준) 0/1
        counts_df.to_excel(writer, sheet_name='LevelCounts')      # 지표×수준 프레임수

    print(f"[완료] 동작 마스크/레벨 분석 저장 → {mask_xlsx}")


    # 여기서 영상 OFF면 바로 종료
    if not MAKE_VIDEO:
        print("[정보] MAKE_VIDEO=False: 시각화 영상 렌더링을 건너뜁니다.")
        return
    
    raw_vals = [[None] * len(pids) for _ in indicators]
    for i, (name, icfg) in enumerate(indicators):
        for j, pid in enumerate(pids):
            df = data_dict[pid]
            if icfg['type'] in ('numeric', 'event'):
                arr = df[icfg['column']].astype(float).values
            else:
                arr = df[icfg['column']].fillna('neutral') \
                        .map(icfg['mapping']).fillna(0).astype(float).values
            if icfg.get('zscore', False):
                m = global_stats[pid][name]['mean']
                s = global_stats[pid][name]['std']
                arr = (arr - m) / s
            raw_vals[i][j] = arr

    fig = plt.figure(figsize=(5, 15))
    gs = GridSpec(1 + len(indicators), len(pids), height_ratios=[1] + [0.7] * len(indicators))
    ims = []
    for idx, pid in enumerate(pids):
        ax = fig.add_subplot(gs[0, idx])
        ax.axis('off')
        ax.set_title(pid)
        frame = caps[pid].read()
        img = (cv2.cvtColor(cv2.resize(frame, (FRAME_WIDTH, FRAME_HEIGHT)), cv2.COLOR_BGR2RGB)
               if frame is not None else np.zeros((FRAME_HEIGHT, FRAME_WIDTH, 3), dtype=np.uint8))
        ims.append(ax.imshow(img))

    plot_axes, line_objs, shade_objs, time_lines = [], [], [], []
    for i, (name, icfg) in enumerate(indicators, 1):
        ax = fig.add_subplot(gs[i, :])
        plot_axes.append(ax)
        # lines for each pid
        lines = [ax.plot([], [], color=colors[j], lw=1.5, zorder=2)[0] for j in range(len(pids))]
        line_objs.append(lines)
        # shading for concurrency levels
        shades = [ax.fill_between([], [], [], color=to_rgba('purple', ((k-1)/(len(pids)-1))*0.6), zorder=0)
                  for k in range(2, len(pids)+1)]
        shade_objs.append(shades)
        # time cursor
        tl = ax.axvline(sf, color='k', ls='--', zorder=2)
        time_lines.append(tl)
        ymin, ymax = icfg.get('ymin'), icfg.get('ymax')
        if ymin is not None and ymax is not None:
            ax.set_ylim(ymin, ymax)
        ax.set_ylabel(name)
        ax.grid(True, zorder=1)
        # numeric/categorical thresholds
        if 'threshold_std' in icfg:
            ax.axhline(icfg['threshold_std'], ls='--', color='gray', zorder=2)
            ax.axhline(-icfg['threshold_std'], ls='--', color='gray', zorder=2)
        # event thresholds
        if icfg.get('type') == 'event':
            ep = icfg.get('event_params', {})
            rt = ep.get('rise_z_thresh')
            ft = ep.get('fall_z_thresh')
            if rt is not None:
                ax.axhline(rt, ls='--', color='gray', zorder=2)
            if ft is not None:
                ax.axhline(ft, ls='--', color='gray', zorder=2)
        if i == 1:
            leg = ax.legend(pids, loc='upper right')
            leg.set_zorder(2)

    plt.tight_layout()
    writer = FFMpegWriter(fps=FPS)
    out_path = os.path.join(timeline_dir, OUTPUT_NAME)
    writer.setup(fig, out_path, dpi=150)

    for f in tqdm(range(sf, ef), desc='Rendering'):
        # update frames
        for idx, pid in enumerate(pids):
            frm = caps[pid].read()
            if frm is not None:
                ims[idx].set_data(
                    cv2.cvtColor(cv2.resize(frm, (FRAME_WIDTH, FRAME_HEIGHT)), cv2.COLOR_BGR2RGB)
                )
        # update plots and shading
        for i, (name, icfg) in enumerate(indicators):
            ax = plot_axes[i]
            start_win = (f // STEP_FRAMES) * STEP_FRAMES
            end_win = min(start_win + STEP_FRAMES, ef)
            x = np.arange(start_win, end_win)
            x_labels = [f"{int(f//(FPS*3600)):02}:{int((f//(FPS*60))%60):02}:{int((f//FPS)%60):02}" for f in x]

            # update lines
            for j, line in enumerate(line_objs[i]):
                line.set_data(x, raw_vals[i][j][start_win:end_win])

            ax.set_xlim(start_win, end_win)

            # tick 간격: 10초 간격
            tick_step = FPS * 10
            tick_indices = np.arange(0, len(x), tick_step)
            ax.set_xticks(x[tick_indices])
            ax.set_xticklabels([x_labels[i] for i in tick_indices], rotation=0)  # ⬅ 수평 출력
            # update shading
            if f % SHADING_FREQ == 0:
                vals = sync_masks[name][start_win:end_win]
                ymin, ymax = ax.get_ylim()
                # remove old shades
                for sh in shade_objs[i]: sh.remove()
                new_shades = []
                for k, sh in enumerate(shade_objs[i], 2):
                    alpha = ((k-1)/(len(pids)-1))*0.6
                    mask_k = vals >= k
                    new_shades.append(
                        ax.fill_between(x, ymin, ymax, where=mask_k, color=to_rgba('purple', alpha), zorder=0)
                    )
                shade_objs[i] = new_shades
        # move time cursors
        for tl in time_lines:
            tl.set_xdata([f, f])
        writer.grab_frame()
    writer.finish()
    print(f"[TIME] Total elapsed: {time.time()-total_start:.2f}s")
    for loader in caps.values(): loader.stop(); loader.cap.release()
    print(f"[완료] 시각화 저장됨 → {out_path}")

if __name__ == '__main__':
    visualize_timeline_optimized(
        timeline_dir="D:/2025신윤희Data/MediaPipe/24-1/A4/W1/T1",
        config_path="config_indicators.json",
        start_time=1500,
        end_time=1620
    )
